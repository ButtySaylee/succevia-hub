import openpyxl, json, sys

wb = openpyxl.load_workbook(r'C:\Users\Hp\Downloads\scholarships_v2 (@timmysofine).xlsx')
ws = wb['Scholarships Database']

headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=3, column=col).value)

print('=== HEADERS ===')
for i, h in enumerate(headers):
    print('  Col {}: {}'.format(i+1, h))

print()
print('=== COUNTING ROWS WITH DATA ===')
count = 0
data = []
for row in range(4, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name and str(name).strip() and not str(name).strip().startswith('*'):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            h = headers[col-1]
            val = ws.cell(row=row, column=col).value
            row_data[h] = str(val).strip() if val else ''
        data.append(row_data)
        count += 1

print('Total scholarships found: {}'.format(count))
print()
print('=== ALL SCHOLARSHIPS ===')
for i, d in enumerate(data):
    name = d.get('Scholarship Name', '?')
    org = d.get('Funder / Organization', '?')
    country = d.get('Host Country', '?')
    level = d.get('Level of Study', '?')
    deadline = d.get('Application Deadline', '?')
    link = d.get('Application Link', '?')
    print('{}. {} | Org: {} | Country: {} | Level: {} | Deadline: {} | Link: {}'.format(i+1, name, org, country, level, deadline, link))

print()
print('Total: {} scholarships'.format(count))

# Save to JSON for later use
with open('scripts/extracted_scholarships.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)
print('Saved to scripts/extracted_scholarships.json')