import openpyxl, json, urllib.request, sys, re

# Step 1: Extract from Excel
print("=" * 60)
print("STEP 1: Extracting scholarships from Excel")
print("=" * 60)

wb = openpyxl.load_workbook(r'C:\Users\Hp\Downloads\scholarships_v2 (@timmysofine).xlsx')
ws = wb['Scholarships Database']

headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=3, column=col).value)

excel_scholarships = []
for row in range(4, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name and str(name).strip() and not str(name).strip().startswith('*'):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            h = headers[col-1]
            val = ws.cell(row=row, column=col).value
            row_data[h] = str(val).strip() if val else ''
        excel_scholarships.append(row_data)

print("Found {} scholarships in Excel".format(len(excel_scholarships)))

# Step 2: Fetch existing scholarships from API
print()
print("=" * 60)
print("STEP 2: Fetching existing scholarships from database")
print("=" * 60)

url = "https://succeviahub.vercel.app/api/opportunities/list?type=scholarship&limit=100"
req = urllib.request.Request(url)
try:
    with urllib.request.urlopen(req) as response:
        data = json.loads(response.read().decode())
        existing = data.get('opportunities', [])
        print("Found {} existing scholarships in database".format(len(existing)))
        
        # Print all existing titles for reference
        for s in existing:
            print("  [DB] {} | Org: {} | URL: {}".format(
                s.get('title', '?'), s.get('organization', '?'), s.get('application_url', '?')))
        
except Exception as e:
    print("ERROR fetching existing scholarships: {}".format(e))
    existing = []

def normalize_title(title):
    """Normalize a title for comparison - remove punctuation, extra spaces, lowercase"""
    t = title.lower().strip()
    # Remove parenthetical content
    t = re.sub(r'\([^)]*\)', '', t)
    # Remove years like 2026, 2027, 2026/2027
    t = re.sub(r'\b20\d{2}\b', '', t)
    t = re.sub(r'\b20\d{2}/\d{2}\b', '', t)
    # Remove punctuation and extra spaces
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def get_significant_words(title):
    """Get significant words from a title (remove common words, keep core terms)"""
    t = normalize_title(title)
    stop_words = {'scholarship', 'scholarships', 'program', 'programme', 'for', 'the', 'and', 'of', 'to', 'in', 'a', 
                  'funded', 'fully', '2026', '2027', '2026/27', '2026/2027'}
    words = t.split()
    return [w for w in words if w not in stop_words and len(w) > 2]

# Step 3: Compare and find new scholarships
print()
print("=" * 60)
print("STEP 3: Comparing and identifying new scholarships")
print("=" * 60)

new_scholarships = []
duplicates = []
uncertain = []

# Build database comparison data
db_titles_normalized = []
db_significant_words = []
for s in existing:
    t = s.get('title', '')
    db_titles_normalized.append(normalize_title(t))
    db_significant_words.append(set(get_significant_words(t)))

for s in excel_scholarships:
    name = s.get('Scholarship Name', '').strip()
    org = s.get('Funder / Organization', '').strip()
    link = s.get('Application Link', '').strip()
    
    if not name:
        continue
    
    name_norm = normalize_title(name)
    name_words = set(get_significant_words(name))
    org_lower = org.lower().strip()
    
    # Check for duplicates
    is_duplicate = False
    match_reason = ""
    best_match_idx = -1
    
    for idx, (db_norm, db_words) in enumerate(zip(db_titles_normalized, db_significant_words)):
        # Check 1: Title overlap (intersection of significant words)
        common_words = name_words & db_words
        if len(common_words) >= 3:
            is_duplicate = True
            match_reason = "Keyword overlap: " + ", ".join(list(common_words)[:4])
            best_match_idx = idx
            break
        
        # Check 2: One title is substring of the other
        if name_norm in db_norm or db_norm in name_norm:
            is_duplicate = True
            match_reason = "Title substring match"
            best_match_idx = idx
            break
    
    if is_duplicate:
        duplicates.append((name, org, match_reason, existing[best_match_idx].get('title', '')))
    elif len(name_words) <= 2:
        # Too few distinguishing words to compare properly
        uncertain.append((name, org, link))
    else:
        new_scholarships.append(s)

print()
print("DUPLICATES (already in database):")
for name, org, reason, db_title in duplicates:
    print("  - [EXCEL] {} ({})".format(name, org))
    print("    Reason: {} ---> [DB] {}".format(reason, db_title))
    print()

print()
print("NEW SCHOLARSHIPS TO IMPORT:")
for s in new_scholarships:
    print("  - {} | Org: {} | Country: {} | Link: {}".format(
        s.get('Scholarship Name', '?'),
        s.get('Funder / Organization', '?'),
        s.get('Host Country', '?'),
        s.get('Application Link', '?')
    ))

if uncertain:
    print()
    print("UNCERTAIN (review manually):")
    for name, org, link in uncertain:
        print("  - {} | Org: {} | Link: {}".format(name, org, link))

print()
print("Summary: {} new, {} duplicates, {} uncertain (out of {} total)".format(
    len(new_scholarships), len(duplicates), len(uncertain), len(excel_scholarships)
))

# Step 4: Generate SQL insert statements
print()
print("=" * 60)
print("STEP 4: Generating SQL insert statements")
print("=" * 60)

if new_scholarships:
    sql_lines = []
    sql_lines.append("-- ============================================================")
    sql_lines.append("-- Insert new scholarships from scholarships_v2.xlsx")
    sql_lines.append("-- Generated: auto")
    sql_lines.append("-- ============================================================")
    sql_lines.append("")
    
    for s in new_scholarships:
        name = s.get('Scholarship Name', '').replace("'", "''")
        org = s.get('Funder / Organization', '').replace("'", "''")
        country = s.get('Host Country', '').replace("'", "''")
        level = s.get('Level of Study', '').replace("'", "''")
        field = s.get('Field of Study', '').replace("'", "''")
        eligible = s.get('Eligible Countries', '').replace("'", "''")
        benefits = s.get('Coverage / Benefits', '').replace("'", "''")
        deadline = s.get('Application Deadline', '').replace("'", "''")
        link = s.get('Application Link', '').replace("'", "''")
        notes = s.get('Notes / Tips', '').replace("'", "''")
        requirements = s.get('Min. Academic Req.', '').replace("'", "''")
        
        # Build description from available fields
        desc_parts = []
        if benefits:
            desc_parts.append("Coverage: {}".format(benefits))
        if level:
            desc_parts.append("Level: {}".format(level))
        if field:
            desc_parts.append("Field: {}".format(field))
        if eligible:
            desc_parts.append("Eligible: {}".format(eligible))
        if notes:
            desc_parts.append(notes)
        
        description = ". ".join(desc_parts)
        if not description:
            description = "Scholarship opportunity at {}.".format(org)
        
        # Build requirements
        req_parts = []
        if requirements:
            req_parts.append(requirements)
        if eligible:
            req_parts.append("Eligible countries: {}".format(eligible))
        
        req_text = "; ".join(req_parts) if req_parts else "Check application link for details"
        
        sql = """INSERT INTO opportunities (title, description, type, organization, location, deadline, requirements, application_url, is_active, is_visible)
VALUES ('{}', '{}', 'scholarship', '{}', '{}', '{}', '{}', '{}', true, true);""".format(
            name, description, org, country, deadline, req_text, link
        )
        sql_lines.append(sql)
    
    sql_text = "\n\n".join(sql_lines)
    
    # Save to file
    with open('scripts/insert-new-scholarships.sql', 'w', encoding='utf-8') as f:
        f.write(sql_text)
    
    print("Generated {} INSERT statements".format(len(new_scholarships)))
    print("Saved to scripts/insert-new-scholarships.sql")
    print()
    print("SQL Preview:")
    print(sql_text[:3000])
else:
    print("No new scholarships to import.")
    with open('scripts/insert-new-scholarships.sql', 'w', encoding='utf-8') as f:
        f.write("-- No new scholarships to import.\n")